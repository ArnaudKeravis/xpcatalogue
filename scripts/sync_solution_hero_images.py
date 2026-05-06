#!/usr/bin/env python3
"""
Sync solution hero images from Sodexo zip + `Solutions images.xlsx`:
  unzip source assets → center-crop resize to a fixed 3:2 frame (matches SolutionHeroTile cover)
  → write `public/images/catalogue/assets/solutions/{id}.png`
  → normalize `heroImage` in `solutionsCatalog.ts` to that path.

Requires: Pillow (pip install Pillow), Python 3.9+
"""

from __future__ import annotations

import argparse
import io
import re
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CATALOG_TS = ROOT / "src/lib/data/solutionsCatalog.ts"
PUBLIC_SOL = ROOT / "public/images/catalogue/assets/solutions"
DEFAULT_XLSX = Path.home() / "Downloads/Solutions images.xlsx"
DEFAULT_ZIP = (
    Path.home()
    / "Library/CloudStorage/OneDrive-SODEXO/Design Community Hub - Documents/"
    "05_Sodexo Labs/Catalogue/Images catalogue/Solutions_Images.zip"
)

# Hero tile ≈ w-72 × h-48 → 3:2; use 2× retina-friendly export.
TARGET_W, TARGET_H = 1200, 800

IMAGE_EXTENSIONS = (
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".WEBP",
    ".PNG",
    ".JPG",
    ".JPEG",
    ".avif",
    ".AVIF",
)

# Excel display name → catalogue id (when norm-match fails).
EXCEL_NAME_TO_ID: dict[str, str] = {
    "Affluence": "affluences",
    "Alberts": "albertsSmoothie",
    "Aquablu IH": "aquablu",
    "Blue Ocean": "blueOceanUvd",
    "Costa Alto machine": "costaCoffeeMachine",
    "Greese to Goodness": "greaseToGoodness",
    "PowerPricing": "pricing",
    "Nando": "relearn",
    "Rego AI": "regoAi",
}

# Excel column “Solutions Images” key → zip stem to use (without extension), when filenames diverge.
IMAGE_KEY_TO_STEM: dict[str, str] = {
    "solution_image_Affluence": "solution_image_Affluences",
    "solution_image_Wando_Analytics": "solution_image_Wando_App",
    "solution_image_Instarinse": "solution_image_Bunzl",
}

# Same hero asset as another row (duplicate brand tiles in catalogue).
DUPLICATE_HERO_FROM: dict[str, str] = {
    # Excel "Sodexo WRX" resolves to the first catalogue row only — mirror asset for Digital Apps entry.
    "sodexoWrx": "sodexoWrxConciergerie",
}


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def parse_catalog(path: Path) -> list[tuple[str, str]]:
    text = path.read_text(encoding="utf-8")
    out: list[tuple[str, str]] = []
    for m in re.finditer(r'id:\s*"([^"]+)"\s*,\s*\n\s*name:\s*"([^"]+)"', text):
        out.append((m.group(1), m.group(2)))
    return out


def resolve_solution_id(excel_name: str, catalog: list[tuple[str, str]]) -> str | None:
    raw = excel_name.strip()
    if raw in EXCEL_NAME_TO_ID:
        return EXCEL_NAME_TO_ID[raw]
    n = norm(raw)
    for sid, sname in catalog:
        if norm(sname) == n:
            return sid
        if norm(sid) == n:
            return sid
    return None


def cover_resize(img: "Image.Image", tw: int, th: int) -> "Image.Image":
    from PIL import Image

    if img.mode in ("RGBA", "P"):
        rgb = Image.new("RGB", img.size, (255, 255, 255))
        if img.mode == "P":
            img = img.convert("RGBA")
        rgb.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
        img = rgb
    elif img.mode != "RGB":
        img = img.convert("RGB")

    iw, ih = img.size
    scale = max(tw / iw, th / ih)
    nw, nh = int(iw * scale), int(ih * scale)
    img = img.resize((nw, nh), Image.Resampling.LANCZOS)
    left = (nw - tw) // 2
    top = (nh - th) // 2
    return img.crop((left, top, left + tw, top + th))


def index_zip(z: zipfile.ZipFile) -> dict[str, str]:
    """norm(full stem including solution_image_…) → zip member path."""
    idx: dict[str, str] = {}
    for name in z.namelist():
        if "__MACOSX" in name or name.endswith("/"):
            continue
        p = Path(name)
        if p.suffix not in IMAGE_EXTENSIONS:
            continue
        stem = p.stem
        idx[norm(stem)] = name
    return idx


def resolve_zip_member(excel_image_key: str, zip_index: dict[str, str]) -> str | None:
    """excel_image_key is e.g. solution_image_4site."""
    if excel_image_key in IMAGE_KEY_TO_STEM:
        excel_image_key = IMAGE_KEY_TO_STEM[excel_image_key]

    def lookup(key: str) -> str | None:
        k = norm(key)
        return zip_index.get(k)

    # Strip parenthetical segments from key (Aquablu (IH) → Aquablu)
    simplified = re.sub(r"\s*\([^)]*\)", "", excel_image_key)
    for candidate in (excel_image_key, simplified):
        hit = lookup(candidate)
        if hit:
            return hit

    return None


def replace_hero_line(ts: str, solution_id: str, line_value: str) -> str:
    anchor = f'id: "{solution_id}"'
    pos = ts.find(anchor)
    if pos == -1:
        return ts
    window = ts[pos : pos + 4500]
    m = re.search(r"^\s*heroImage:\s*[^\n]+", window, re.MULTILINE)
    if not m:
        return ts
    abs_start = pos + m.start()
    abs_end = pos + m.end()
    indent = re.match(r"^(\s*)", m.group(0)).group(1)
    new_line = f'{indent}heroImage: "{line_value}",'
    return ts[:abs_start] + new_line + ts[abs_end:]


def main() -> None:
    import sys

    try:
        from PIL import Image
    except ImportError:
        print("Install Pillow: pip install Pillow", file=sys.stderr)
        sys.exit(1)

    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--zip", type=Path, default=DEFAULT_ZIP)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.xlsx.is_file():
        print(f"Missing Excel: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    if not args.zip.is_file():
        print(f"Missing zip: {args.zip}", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    catalog = parse_catalog(CATALOG_TS)
    cat_ids = {sid for sid, _ in catalog}

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    if "Solutions" not in wb.sheetnames:
        print("No 'Solutions' sheet", file=sys.stderr)
        sys.exit(1)
    ws = wb["Solutions"]

    PUBLIC_SOL.mkdir(parents=True, exist_ok=True)

    z = zipfile.ZipFile(args.zip)
    zip_index = index_zip(z)

    updated_ids: list[str] = []
    skipped: list[tuple[str, str]] = []

    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = row[0]
        img_key = row[1]
        if not name or not img_key:
            continue
        sid = resolve_solution_id(str(name), catalog)
        if not sid or sid not in cat_ids:
            skipped.append((str(name), "no catalogue id"))
            continue

        member = resolve_zip_member(str(img_key).strip(), zip_index)
        if not member:
            skipped.append((str(name), f"no zip file for {img_key}"))
            continue

        raw = z.read(member)
        img = Image.open(io.BytesIO(raw))
        img = cover_resize(img, TARGET_W, TARGET_H)

        out_path = PUBLIC_SOL / f"{sid}.png"
        if not args.dry_run:
            img.save(out_path, "PNG", optimize=True)
        updated_ids.append(sid)
        print(f"OK {sid} ← {member}")

    z.close()

    if not args.dry_run:
        import shutil

        for target, source in DUPLICATE_HERO_FROM.items():
            src = PUBLIC_SOL / f"{source}.png"
            dst = PUBLIC_SOL / f"{target}.png"
            if src.is_file():
                shutil.copy2(src, dst)
                updated_ids.append(target)
                print(f"OK {target} ← copy({source}.png)")

    if not args.dry_run and updated_ids:
        ts_text = CATALOG_TS.read_text(encoding="utf-8")
        hero_path_prefix = "/images/catalogue/assets/solutions/"
        for sid in sorted(set(updated_ids)):
            full = f"{hero_path_prefix}{sid}.png"
            ts_text = replace_hero_line(ts_text, sid, full)
        CATALOG_TS.write_text(ts_text, encoding="utf-8")
        print(f"Updated heroImage for {len(set(updated_ids))} solutions in solutionsCatalog.ts")

    if skipped:
        print("\nSkipped:", len(skipped))
        for a, b in skipped[:25]:
            print(f"  {a}: {b}")
        if len(skipped) > 25:
            print("  ...")


if __name__ == "__main__":
    main()
