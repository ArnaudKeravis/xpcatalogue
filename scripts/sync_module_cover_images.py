#!/usr/bin/env python3
"""
Sync module cover images from Sodexo `Modules_Images` + `Modules.xlsx`:
  resize to the catalogue module frame (≈2.31:1) → write
  `public/images/catalogue/modules/{slug}.png` where slug matches
  `moduleCoverImageSlug()` in modulesExcelMerge.ts.

Requires: Pillow, openpyxl (pip install Pillow openpyxl)
"""

from __future__ import annotations

import argparse
import io
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PUBLIC_MOD = ROOT / "public/images/catalogue/modules"
DEFAULT_XLSX = Path.home() / "Downloads/Classeur Modules.xlsx"
DEFAULT_SRC = (
    Path.home()
    / "Library/CloudStorage/OneDrive-SODEXO/Design Community Hub - Documents/"
    "05_Sodexo Labs/Catalogue/Images catalogue/Modules_Images"
)

# Source art is 4620×2000 (46 files) — export 2× retina-friendly frame.
TARGET_W, TARGET_H = 924, 400

IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp", ".PNG", ".JPG", ".JPEG", ".WEBP")


def module_cover_image_slug(image_key: str) -> str:
    """Mirror of moduleCoverImageSlug() in modulesExcelMerge.ts."""
    tail = re.sub(r"^module_image_", "", image_key.strip(), flags=re.I).strip()
    slug = tail.lower().replace("&", "and")
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")


def fit_resize(img: "Image.Image", tw: int, th: int) -> "Image.Image":
    from PIL import Image

    if img.mode in ("RGBA", "P"):
        rgb = Image.new("RGB", img.size, (255, 255, 255))
        if img.mode == "P":
            img = img.convert("RGBA")
        rgb.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
        img = rgb
    elif img.mode != "RGB":
        img = img.convert("RGB")

    return img.resize((tw, th), Image.Resampling.LANCZOS)


def resolve_source_file(src_dir: Path, image_key: str) -> Path | None:
    direct = src_dir / f"{image_key}.png"
    if direct.is_file():
        return direct
    key_norm = re.sub(r"[^a-z0-9]", "", image_key.lower())
    for p in src_dir.iterdir():
        if not p.is_file() or p.suffix.lower() not in {e.lower() for e in IMAGE_EXTENSIONS}:
            continue
        if re.sub(r"[^a-z0-9]", "", p.stem.lower()) == key_norm:
            return p
    return None


def main() -> None:
    try:
        from PIL import Image
    except ImportError:
        print("Install Pillow: pip install Pillow", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--src", type=Path, default=DEFAULT_SRC)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.xlsx.is_file():
        print(f"Missing Excel: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    if not args.src.is_dir():
        print(f"Missing source dir: {args.src}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    if "Modules" not in wb.sheetnames:
        print("No 'Modules' sheet", file=sys.stderr)
        sys.exit(1)
    ws = wb["Modules"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx_img = header.index("Module Image") if "Module Image" in header else 3
    idx_name = header.index("Name of the Module") if "Name of the Module" in header else 0

    PUBLIC_MOD.mkdir(parents=True, exist_ok=True)

    ok: list[str] = []
    skipped: list[tuple[str, str]] = []

    for row in rows[1:]:
        name = row[idx_name] if idx_name < len(row) else None
        img_key = row[idx_img] if idx_img < len(row) else None
        if not name or not img_key:
            continue
        image_key = str(img_key).strip()
        slug = module_cover_image_slug(image_key)
        src_file = resolve_source_file(args.src, image_key)
        if not src_file:
            skipped.append((str(name), image_key))
            continue

        out_path = PUBLIC_MOD / f"{slug}.png"
        if args.dry_run:
            ok.append(slug)
            print(f"DRY {slug} ← {src_file.name}")
            continue

        img = Image.open(src_file)
        img = fit_resize(img, TARGET_W, TARGET_H)
        img.save(out_path, "PNG", optimize=True)
        ok.append(slug)
        print(f"OK {slug} ← {src_file.name}")

    print(f"\nSynced {len(ok)} module covers → {PUBLIC_MOD}")
    if skipped:
        print(f"Skipped {len(skipped)} (no source file — UI falls back to icon):")
        for n, k in skipped:
            print(f"  {n}: {k}")


if __name__ == "__main__":
    main()
