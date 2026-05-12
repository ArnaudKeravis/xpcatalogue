#!/usr/bin/env python3
"""Read **Personae Journey** (Classeur / Moments workbook), Column **D** (Personae image
nomenclature), copy the matching raster from the design asset folder into `public/`, and emit
`momentPersonaTop.generated.ts`.

Column D in the current workbook is **Personae Image Portrait** (filename stem, e.g.
`personae_WORK_portfolio_manager_portrait`). Assets must exist under `--assets-dir` with that
stem + extension (`.png` / `.jpg`). Default `--assets-dir` is `Personae_Images_Portrait` next to
the catalogue Images folder — this is the folder whose filenames align with Column D.

The separate `Personae_Images_Face.zip` / Face folder uses `*_face` stems (Column C); it does
not contain Column D portrait files. Override `--assets-dir` only if your Column D files live
elsewhere.

Row → catalogue persona + journey step uses the same pairing / title resolution as
`generate-moment-excel-assets.py`.

Usage:
  python3 scripts/generate_moment_persona_top_from_excel.py \\
    --xlsx ~/Downloads/Classeur\\ Journey.xlsx \\
    [--assets-dir \".../Personae_Images_Portrait\"]
"""

from __future__ import annotations

import argparse
import json
import platform
import re
import shutil
import subprocess
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise

REPO = Path(__file__).resolve().parent.parent

PAIR_TO_ID: dict[tuple[str, str], str] = {
    ("work", "portfolio manager"): "client-work",
    ("work", "site manager"): "operator-work",
    ("work", "production line worker"): "blue-collar",
    ("work", "lab researcher"): "grey-collar",
    ("work", "army officer"): "military",
    ("work", "admin function"): "white-collar",
    ("heal", "portfolio manager"): "client-heal",
    ("heal", "site manager"): "operator-heal",
    ("heal", "hospital physician"): "doctor",
    ("heal", "registered nurse"): "nurse",
    ("heal", "long-term resident"): "senior",
    ("heal", "outpatient & inpatient"): "patient",
    ("learn", "portfolio manager"): "client-learn",
    ("learn", "site manager"): "operator-learn",
    ("learn", "university student"): "student",
    ("learn", "middle school student"): "schoolchild",
    ("learn", "middle school teacher"): "teacher",
    ("learn", "working parent of school-age children"): "parent",
    ("play", "portfolio manager"): "client-play",
    ("play", "site manager"): "operator-play",
    ("play", "business traveller & vip guest"): "tourist",
    ("play", "cultural tourist & museum visitor"): "tourist",
    ("play", "executive & premium event guest"): "vip-guest",
    ("play", "parent & sports fan"): "sport-fan",
    ("play", "trade show & event attendent"): "participant",
}

TITLE_TO_STEP: dict[tuple[str, str], str] = {
    ("tourist", "arrival & access"): "tou-welcome",
    ("tourist", "social time/work"): "tou-visit",
    ("tourist", "fast access to gate"): "tou-depart",
    ("schoolchild", "welcome at school"): "sch-welcome",
}


def norm_persona(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s).replace("\n", " ")).strip()


def norm_title(t: Any) -> str:
    s = re.sub(r"\s+", " ", str(t).lower().strip())
    s = s.replace("kick-off", "kick off")
    s = re.sub(r"\s*/\s*", "/", s)
    s = s.replace("social time / work", "social time/work")
    return s


def sanitize_key(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).replace("\n", "").replace("\r", "").strip()


def norm_file_stem(stem: str) -> str:
    s = stem.lower()
    s = re.sub(r"_+", "_", s)
    return s


def load_moment_labels(repo: Path) -> dict[str, list[dict[str, str]]]:
    jpath = repo / "scripts/persona-moments.json"
    if not jpath.exists():
        raise SystemExit(f"missing {jpath}")
    return json.loads(jpath.read_text(encoding="utf-8"))


def resolve_pair(area_raw: Any, persona_raw: Any) -> str | None:
    area = str(area_raw).strip().lower()
    pers = norm_persona(persona_raw).lower()
    return PAIR_TO_ID.get((area, pers))


def include_row_for_persona(persona_id: str, excel_persona_raw: Any) -> bool:
    x = norm_persona(excel_persona_raw).lower()
    if persona_id == "tourist":
        return "business traveller" in x and "cultural tourist" not in x
    return True


def resolve_step_id(
    persona_id: str,
    excel_title: Any,
    moments: list[dict[str, str]],
) -> tuple[str | None, float]:
    alias = TITLE_TO_STEP.get((persona_id, norm_title(excel_title)))
    if alias:
        return alias, 1.0
    nt = norm_title(excel_title)
    for m in moments:
        if norm_title(m["label"]) == nt:
            return m["id"], 1.0
    best_id: str | None = None
    best = 0.0
    for m in moments:
        sc = SequenceMatcher(None, nt, norm_title(m["label"])).ratio()
        if sc > best:
            best = sc
            best_id = m["id"]
    return best_id, best


def header_map(ws: Any) -> dict[str, int]:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    out: dict[str, int] = {}
    for i, cell in enumerate(row1):
        if cell is None:
            continue
        key = re.sub(r"\s+", " ", str(cell).strip().lower())
        out[key] = i
    return out


def col_moment_title(h: dict[str, int]) -> int:
    for k, idx in h.items():
        if "moments title" in k or k == "moments title":
            return idx
    return 5


def col_personae_image_d(h: dict[str, int]) -> int:
    """Excel column D (0-based index 3) — Personae Image Portrait in Classeur Journey."""
    for label in ("personae image portrait", "personae image portrait "):
        if label in h:
            return h[label]
    return 3


def build_index_from_dir(images_dir: Path) -> dict[str, Path]:
    index: dict[str, Path] = {}
    for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
        for p in images_dir.glob(ext):
            index[norm_file_stem(p.stem)] = p
    return index


def shrink_for_web_tile(path: Path, max_edge: int = 512) -> None:
    """Downscale rasters for horizontal moment cards (keeps repo / deploy size reasonable)."""
    if platform.system() != "Darwin":
        return
    if path.suffix.lower() not in (".png", ".jpg", ".jpeg"):
        return
    try:
        subprocess.run(
            ["sips", "-Z", str(max_edge), str(path), "--out", str(path)],
            check=True,
            capture_output=True,
        )
    except (FileNotFoundError, subprocess.CalledProcessError):
        pass


def emit_ts(data: dict[str, dict[str, str]]) -> str:
    lines = [
        "/** Auto-generated by scripts/generate_moment_persona_top_from_excel.py — do not edit by hand.",
        " *  Maps catalogue persona id + journey step id → public URL for the persona image at the top",
        " *  of moment cards (Column D / Personae Image Portrait in Personae Journey). */",
        "",
        "export const MOMENT_PERSONA_TOP: Partial<",
        "  Record<string, Partial<Record<string, string>>>",
        "> = {",
    ]
    for pid in sorted(data):
        lines.append(f'  "{pid}": {{')
        for sid in sorted(data[pid]):
            url = data[pid][sid].replace("\\", "\\\\").replace('"', '\\"')
            lines.append(f'    "{sid}": "{url}",')
        lines.append("  },")
    lines.append("};")
    lines.append("")
    lines.append(
        "export function momentPersonaTopUrl(personaId: string, stepId: string): string | undefined {"
    )
    lines.append("  return MOMENT_PERSONA_TOP[personaId]?.[stepId];")
    lines.append("}")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path.home() / "Downloads/Classeur Journey.xlsx")
    ap.add_argument(
        "--assets-dir",
        type=Path,
        default=None,
        help="Folder of portrait images whose stems match Column D (default: Personae_Images_Portrait)",
    )
    args = ap.parse_args()

    default_portrait = (
        Path.home()
        / "Library/CloudStorage/OneDrive-SODEXO/Design Community Hub - Documents"
        / "05_Sodexo Labs/Catalogue/Images catalogue/Personae_Images_Portrait"
    )
    assets_dir = args.assets_dir or default_portrait

    if not args.xlsx.exists():
        print(f"error: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 2
    if not assets_dir.is_dir():
        print(f"error: assets dir not found: {assets_dir}", file=sys.stderr)
        return 2

    labels = load_moment_labels(REPO)
    img_index = build_index_from_dir(assets_dir)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    if "Personae Journey" not in wb.sheetnames:
        print("error: sheet 'Personae Journey' missing", file=sys.stderr)
        return 2
    ws = wb["Personae Journey"]
    h = header_map(ws)
    idx_title = col_moment_title(h)
    idx_col_d = col_personae_image_d(h)

    out: dict[str, dict[str, str]] = {}
    out_dir = REPO / "public/images/catalogue/assets/journeys/moment-persona-top"
    missing: list[tuple[str, str, str]] = []
    weak: list[tuple[str, str, str | None, float]] = []
    skipped_pair: list[tuple[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        pid = resolve_pair(row[0], row[1])
        if not pid:
            skipped_pair.append((str(row[0]), norm_persona(row[1])))
            continue
        if not include_row_for_persona(pid, row[1]):
            continue
        title = row[idx_title] if idx_title < len(row) else None
        col_d = row[idx_col_d] if idx_col_d < len(row) else None
        ref = sanitize_key(col_d)
        if not ref:
            missing.append((pid, str(title or ""), "<empty Column D>"))
            continue

        moments = labels.get(pid, [])
        step_id, score = resolve_step_id(pid, title, moments)
        if not step_id:
            weak.append((pid, str(title), None, score))
            continue
        if score < 0.86 and (pid, norm_title(title)) not in TITLE_TO_STEP:
            weak.append((pid, str(title), step_id, score))

        stem = norm_file_stem(Path(ref).stem if "." in ref else ref)
        src = img_index.get(stem)
        if src is None:
            raw_name = sanitize_key(ref)
            if "/" in raw_name:
                raw_name = raw_name.split("/")[-1]
            if raw_name.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                cand = assets_dir / raw_name
                if cand.is_file():
                    src = cand
        if src is None:
            missing.append((pid, step_id, ref))
            continue

        dest_sub = out_dir / pid
        dest_sub.mkdir(parents=True, exist_ok=True)
        ext = src.suffix.lower() or ".png"
        dest = dest_sub / f"{step_id}{ext}"
        shutil.copyfile(src, dest)
        shrink_for_web_tile(dest)
        rel = f"/images/catalogue/assets/journeys/moment-persona-top/{pid}/{step_id}{ext}"
        out.setdefault(pid, {})[step_id] = rel

    out_ts = REPO / "src/lib/data/momentPersonaTop.generated.ts"
    out_ts.write_text(emit_ts(out), encoding="utf-8")

    print(f"wrote {out_ts.relative_to(REPO)} — {sum(len(v) for v in out.values())} persona×moment URLs")
    print(f"copied rasters under {out_dir.relative_to(REPO)}")
    if skipped_pair:
        print(f"WARNING: unknown area/persona pairs: {len(skipped_pair)}")
        for p in skipped_pair[:8]:
            print(" ", p)
    if weak:
        print(f"WARNING: weak moment title matches: {len(weak)}")
        for w in weak[:10]:
            print(" ", w)
    if missing:
        print(f"ERROR: missing Column D asset or empty ref: {len(missing)}")
        for m in missing[:20]:
            print(" ", m)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
