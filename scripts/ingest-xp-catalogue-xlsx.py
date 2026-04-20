#!/usr/bin/env python3
"""Re-ingest the authoritative XP Catalogue flow from Excel into TypeScript.

Source:  /Users/<you>/Downloads/Catalogue_XP_solutions.xlsx
         (override with --input)
Target:  src/lib/data/xpCatalogueFlow.ts

Produces a normalized structure consumed by `src/lib/data/xpFlowAdapter.ts`:
  - modules              (55 existing + 7 new suggested)
  - consumerWorkMoments  (White-Collar consumer day)
  - operatorMoments      (Operator day)
  - solutions            (~108 with description + mapping flags)

The adapter handles fuzzy matching between Excel solution names and app
`SOLUTIONS_CATALOG` ids, so this file just preserves the Excel verbatim.

Run:  python3 scripts/ingest-xp-catalogue-xlsx.py
      python3 scripts/ingest-xp-catalogue-xlsx.py --input /path/to/other.xlsx
      python3 scripts/ingest-xp-catalogue-xlsx.py --check   # exit 1 if stale
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "error: openpyxl missing. Install via: pip install openpyxl\n"
    )
    sys.exit(2)


REPO = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path.home() / "Downloads" / "Catalogue_XP_solutions.xlsx"
TARGET_TS = REPO / "src" / "lib" / "data" / "xpCatalogueFlow.ts"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def clean(v) -> str:
    """Strip + collapse internal whitespace. Returns '' for None."""
    if v is None:
        return ""
    s = str(v).strip()
    # Collapse double spaces but preserve intentional punctuation
    s = re.sub(r"\s+", " ", s)
    return s


def slugify(name: str, persona: str | None = None) -> str:
    base = name.lower()
    base = base.replace("&", "and")
    base = re.sub(r"\(nouveau\)", "", base).strip()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    if persona:
        base = f"{base}-{persona.lower()}"
    return base


def is_new(statut: str, name: str) -> bool:
    """Module is 'new' when statut contains 🆕 NOUVEAU or name tagged (NOUVEAU)."""
    if statut and "NOUVEAU" in statut.upper():
        return True
    if "NOUVEAU" in name.upper():
        return True
    return False


def strip_new_suffix(name: str) -> str:
    return re.sub(r"\s*\(NOUVEAU\)\s*$", "", name, flags=re.IGNORECASE).strip()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet parsers
# ─────────────────────────────────────────────────────────────────────────────

def parse_modules(ws) -> list[dict]:
    """Parse the 'Modules' sheet → [{slug, name, persona, description, solutions, isNew}].

    Layout from the current file (Mar 2026):
      col0: (blank)
      col1: Module name
      col2: Persona (Consumer|Operator)
      col3: Description
      col4: 'Solutions associées' (comma-separated)
      col5: Statut ('✅ Existant' | '🆕 NOUVEAU')
    Header rows occupy rows 1-3; data starts at row 4.
    """
    modules: list[dict] = []
    seen_slugs: set[str] = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = clean(row[1])
        if not name:
            continue
        persona = clean(row[2]) or None
        description = clean(row[3])
        solutions_raw = clean(row[4])
        statut = clean(row[5]) if len(row) > 5 else ""

        new_flag = is_new(statut, name)
        display_name = strip_new_suffix(name)

        solutions = [s.strip() for s in re.split(r",\s*", solutions_raw) if s.strip()]

        slug = slugify(display_name, persona)
        # Dedup identical (name,persona) by suffixing an index
        if slug in seen_slugs:
            idx = 2
            while f"{slug}-{idx}" in seen_slugs:
                idx += 1
            slug = f"{slug}-{idx}"
        seen_slugs.add(slug)

        modules.append({
            "slug": slug,
            "name": display_name,
            "persona": persona,
            "description": description,
            "solutions": solutions,
            "isNew": new_flag,
        })
    return modules


def parse_consumer_wk(ws) -> list[dict]:
    """Parse 'Consumer WK' sheet into [{name, modules: [{module, solutions}]}].

    Layout: col0 is moment header ('Commute:', 'Welcome Area', 'F&B ', 'WP',
    'Wellbeing'). col1 is module name. col2+ are solutions for that module.
    Header row is row 1; data starts at row 2. Moment headers carry trailing ':'.
    """
    moments: list[dict] = []
    current: dict | None = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        moment_cell = clean(row[0])
        module_cell = clean(row[1])
        solution_cells = [clean(c) for c in row[2:] if clean(c)]

        if moment_cell:
            # New moment block
            moment_name = moment_cell.rstrip(":")
            current = {"name": moment_name, "modules": []}
            moments.append(current)
            # Module on the same row as the header? (rare but possible)
            if module_cell:
                current["modules"].append({
                    "module": module_cell,
                    "solutions": solution_cells,
                })
            continue

        if not current:
            continue  # stray row before any moment header

        if module_cell:
            current["modules"].append({
                "module": module_cell,
                "solutions": solution_cells,
            })
    # Drop empty moments
    return [m for m in moments if m["modules"]]


def parse_operator(ws) -> list[dict]:
    """Parse 'Operator' sheet into [{name, modules: [{module, solutions}]}].

    Layout: col0 is persona ('Operator', first data row only). col1 is moment
    name (sparse — carries forward). col2 is module name. col3+ are solutions.
    """
    moments: list[dict] = []
    current: dict | None = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        moment_cell = clean(row[1]) if len(row) > 1 else ""
        module_cell = clean(row[2]) if len(row) > 2 else ""
        solution_cells = [clean(c) for c in row[3:] if clean(c)]

        if moment_cell:
            current = {"name": moment_cell, "modules": []}
            moments.append(current)

        if not current:
            continue

        if module_cell:
            current["modules"].append({
                "module": module_cell,
                "solutions": solution_cells,
            })
    return [m for m in moments if m["modules"]]


def parse_solutions(ws) -> dict[str, dict]:
    """Parse 'Solutions' sheet → {name: {description, mappedConsumerWC, mappedOperator}}."""
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 5:
            continue
        _idx, name, desc, m_consumer, m_operator = row[:5]
        name = clean(name)
        if not name:
            continue
        out[name] = {
            "description": clean(desc),
            "mappedConsumerWC": clean(m_consumer).startswith("✓"),
            "mappedOperator": clean(m_operator).startswith("✓"),
        }
    return out


# ─────────────────────────────────────────────────────────────────────────────
# TypeScript emission
# ─────────────────────────────────────────────────────────────────────────────

TS_HEADER = '''/**
 * XP Catalogue Flow — authoritative data from `Catalogue_XP_solutions.xlsx`.
 *
 * GENERATED FILE — edit the source Excel and run:
 *   python3 scripts/ingest-xp-catalogue-xlsx.py
 *
 * Structure:
 *   - `modules`: 55 modules + 7 new proposed (Consumer + Operator personas)
 *   - `consumerWorkMoments`: White-collar consumer journey moments → modules → solutions
 *   - `operatorMoments`: Operator journey moments → modules → solutions
 *   - `solutions`: All solutions with Excel descriptions + mapping flags
 */

export interface XpFlowModule {
  /** URL-safe slug derived from module name + persona */
  slug: string;
  /** Display name */
  name: string;
  /** Who owns this module journey — 'Consumer' or 'Operator' */
  persona: string | null;
  /** Description sourced from the Excel Modules tab */
  description: string;
  /** Solution names (Excel spelling) mapped to this module */
  solutions: string[];
  /** True when the module is marked as new/proposed (orange) */
  isNew: boolean;
}

export interface XpFlowMomentModule {
  module: string;
  solutions: string[];
}

export interface XpFlowMoment {
  name: string;
  modules: XpFlowMomentModule[];
}

export interface XpFlowSolutionMeta {
  description: string;
  mappedConsumerWC: boolean;
  mappedOperator: boolean;
}

export interface XpCatalogueFlow {
  modules: XpFlowModule[];
  consumerWorkMoments: XpFlowMoment[];
  operatorMoments: XpFlowMoment[];
  solutions: Record<string, XpFlowSolutionMeta>;
}

export const XP_CATALOGUE_FLOW: XpCatalogueFlow = '''


def emit_ts(data: dict) -> str:
    """Pretty-print the data structure as a JSON5-friendly TS literal.

    Using JSON is fine because TS accepts double-quoted object literals.
    """
    # json.dumps preserves insertion order in py3.7+, guaranteeing deterministic diff.
    body = json.dumps(data, indent=2, ensure_ascii=False)
    return TS_HEADER + body + ";\n"


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="Regenerate xpCatalogueFlow.ts from Excel")
    parser.add_argument("--input", type=Path, default=DEFAULT_XLSX, help="Source .xlsx (default: ~/Downloads/Catalogue_XP_solutions.xlsx)")
    parser.add_argument("--output", type=Path, default=TARGET_TS, help="Target .ts (default: src/lib/data/xpCatalogueFlow.ts)")
    parser.add_argument("--check", action="store_true", help="Exit 1 if output differs from generated content (for CI)")
    parser.add_argument("--quiet", action="store_true", help="Suppress diff summary")
    args = parser.parse_args()

    xlsx: Path = args.input
    out: Path = args.output

    if not xlsx.exists():
        sys.stderr.write(f"error: input not found: {xlsx}\n")
        return 2

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    # Sheet names may have trailing whitespace. Normalize lookup.
    sheet_map = {name.strip(): wb[name] for name in wb.sheetnames}
    required = ["Modules", "Consumer WK", "Operator", "Solutions"]
    for req in required:
        if req not in sheet_map:
            sys.stderr.write(f"error: sheet '{req}' missing. Found: {list(sheet_map)}\n")
            return 2

    data = {
        "modules": parse_modules(sheet_map["Modules"]),
        "consumerWorkMoments": parse_consumer_wk(sheet_map["Consumer WK"]),
        "operatorMoments": parse_operator(sheet_map["Operator"]),
        "solutions": parse_solutions(sheet_map["Solutions"]),
    }

    # Sanity report
    if not args.quiet:
        print(f"  modules:             {len(data['modules']):>4d}  ({sum(1 for m in data['modules'] if m['isNew'])} new)")
        print(f"  consumerWorkMoments: {len(data['consumerWorkMoments']):>4d}")
        print(f"  operatorMoments:     {len(data['operatorMoments']):>4d}")
        print(f"  solutions:           {len(data['solutions']):>4d}")

    new_content = emit_ts(data)
    old_content = out.read_text(encoding="utf-8") if out.exists() else ""

    if new_content == old_content:
        if not args.quiet:
            print(f"\n✓ {out.relative_to(REPO)} already up-to-date")
        return 0

    if args.check:
        sys.stderr.write(f"\n✗ {out.relative_to(REPO)} is stale — re-run without --check\n")
        return 1

    out.write_text(new_content, encoding="utf-8")
    if not args.quiet:
        print(f"\n✓ wrote {out.relative_to(REPO)} ({len(new_content):,} chars)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
