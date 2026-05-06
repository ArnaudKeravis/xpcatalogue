#!/usr/bin/env python3
"""Emit src/lib/data/modulesExcelSoT.generated.ts from Modules.xlsx (Sodexo catalogue SoT)."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("openpyxl is required: pip install openpyxl") from e

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src/lib/data/modulesExcelSoT.generated.ts"
DEFAULT_XLSX = Path.home() / "Downloads/Modules.xlsx"


def esc(s: str) -> str:
    return json.dumps(s, ensure_ascii=True)


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument(
        "xlsx",
        nargs="?",
        default=str(DEFAULT_XLSX),
        help="Path to Modules.xlsx",
    )
    args = p.parse_args()
    path = Path(args.xlsx).expanduser()
    if not path.is_file():
        raise SystemExit(f"Missing workbook: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Modules" not in wb.sheetnames:
        raise SystemExit(f"No 'Modules' sheet in {path}")
    ws = wb["Modules"]
    rows_in = list(ws.iter_rows(values_only=True))
    if not rows_in:
        raise SystemExit("Empty sheet")
    header = [str(c).strip() if c is not None else "" for c in rows_in[0]]
    # Expected columns from Sodexo template
    idx_name = header.index("Name of the Module") if "Name of the Module" in header else 0
    idx_domain = header.index("Domain / Macro Personae") if "Domain / Macro Personae" in header else 1
    idx_sol = header.index("Solutions in the module") if "Solutions in the module" in header else 2
    idx_img = header.index("Module Image") if "Module Image" in header else 3
    idx_desc = header.index("Module Description") if "Module Description" in header else 4
    idx_frame = header.index("Module Frame") if "Module Frame" in header else -1

    rows: list[dict] = []
    for row in rows_in[1:]:
        name = row[idx_name] if idx_name < len(row) else None
        if name is None or str(name).strip() == "":
            continue
        domain = row[idx_domain] if idx_domain < len(row) else ""
        sol = row[idx_sol] if idx_sol < len(row) else ""
        img = row[idx_img] if idx_img < len(row) else ""
        desc = row[idx_desc] if idx_desc < len(row) else ""
        frame = row[idx_frame] if idx_frame >= 0 and idx_frame < len(row) else None
        rows.append(
            {
                "name": str(name).strip(),
                "domain": str(domain).strip() if domain else "",
                "solutionNamesRaw": str(sol).strip() if sol else "",
                "imageKey": str(img).strip() if img else "",
                "description": str(desc).strip() if desc else "",
                "frame": str(frame).strip() if frame else None,
            }
        )

    lines = [
        "/**",
        " * Auto-generated from Sodexo `Modules.xlsx` — do not edit by hand.",
        f" * Source rows: {len(rows)}",
        " */",
        "",
        "export interface ModulesExcelRow {",
        "  name: string;",
        "  domain: string;",
        "  solutionNamesRaw: string;",
        "  imageKey: string;",
        "  description: string;",
        "  frame: string | null;",
        "}",
        "",
        "export const MODULES_EXCEL_SOT: readonly ModulesExcelRow[] = [",
    ]
    for r in rows:
        frame_lit = "null" if r["frame"] is None else esc(r["frame"])
        lines.append(
            "  {"
            f"name: {esc(r['name'])}, "
            f"domain: {esc(r['domain'])}, "
            f"solutionNamesRaw: {esc(r['solutionNamesRaw'])}, "
            f"imageKey: {esc(r['imageKey'])}, "
            f"description: {esc(r['description'])}, "
            f"frame: {frame_lit}"
            "},"
        )
    lines.append("];")
    lines.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT} ({len(rows)} modules)")


if __name__ == "__main__":
    main()
