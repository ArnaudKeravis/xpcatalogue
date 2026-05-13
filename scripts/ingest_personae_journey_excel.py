#!/usr/bin/env python3
"""Rebuild persona journeys from **Personae Journey** (Classeur workbook).

Uses header names (not fixed letters) so column shifts are tolerated. Canonical columns:

  - **Areas** (Excel A) · **Personae** (B) · **Personae Name** (C) — identify the catalogue persona row
  - **Personae Iso Journey** (G) — snake_case key naming the **journey map** asset; must match
    ``ISO_JOURNEY_KEY_TO_FILENAME`` and a file under ``--journey-assets`` (recursive search).
  - **Moments Title** (H) — moment label; row order = journey order
  - **Image left moment** (K) — sole source for the **moment image** asset key; resolved on disk
    under Journey_Moments_Images, then copied to ``public/.../moments-raster/{personaId}/{stepId}.*``
    and emitted as ``MOMENT_HERO_RASTER`` (used by journey moment cards + moment page hero).

Other columns (subtitle, body, modules, portrait) are unchanged.

Row order in the sheet = journey order per persona. Column **Modules** is split on comma/semicolon;
each token is trimmed and stored verbatim on the `JourneyStep`.

Emits:
  - src/lib/data/personaeJourneyExcel.generated.ts
  - src/lib/data/journeyStepsFromExcel.generated.ts
  - src/lib/data/momentEditorial.generated.ts
  - src/lib/data/momentHeroRaster.generated.ts
  - src/lib/data/momentPersonaTop.generated.ts
  - scripts/persona-moments.json
  - copies under public/ for hero + persona-top rasters

Usage:
  python3 scripts/ingest_personae_journey_excel.py \\
    --xlsx ~/Downloads/Classeur\\ Journey.xlsx \\
    --journey-assets ~/Downloads

Requires ``numpy`` and ``pillow`` for marker detection. Rasterization: macOS ``qlmanage``,
else ``magick`` or ``rsvg-convert`` for SVG.
"""

from __future__ import annotations

import argparse
import json
import platform
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

try:
    from journey_marker_geometry import hotspot_boxes_for_moments, rasterize_journey_asset
except ImportError:
    print("pip install numpy pillow", file=sys.stderr)
    raise

REPO = Path(__file__).resolve().parent.parent

PAIR_TO_ID: dict[tuple[str, str], str] = {
    ("work", "portfolio manager"): "client-work",
    ("work", "site manager"): "operator-work",
    ("work", "production line worker"): "blue-collar",
    ("work", "lab researcher"): "grey-collar",
    ("work", "army officer"): "military",
    ("work", "admin function"): "white-collar",
    ("heal", "portfolio manager"): "client-heal",
    ("heal", "site manager"): "operator-heal",
    ("heal", "hospital physician"): "doctor",
    ("heal", "registered nurse"): "nurse",
    ("heal", "long-term resident"): "senior",
    ("heal", "outpatient & inpatient"): "patient",
    ("learn", "portfolio manager"): "client-learn",
    ("learn", "site manager"): "operator-learn",
    ("learn", "university student"): "student",
    ("learn", "middle school student"): "schoolchild",
    ("learn", "middle school teacher"): "teacher",
    ("learn", "working parent of school-age children"): "parent",
    ("play", "portfolio manager"): "client-play",
    ("play", "site manager"): "operator-play",
    ("play", "business traveller & vip guest"): "tourist",
    ("play", "cultural tourist & museum visitor"): "tourist",
    ("play", "executive & premium event guest"): "vip-guest",
    ("play", "parent & sports fan"): "sport-fan",
    ("play", "trade show & event attendent"): "participant",
}

# Exact **Personae Iso Journey** cell values (Column G) from Classeur → journey artwork filename.
# Whitespace is stripped for lookup; see ``norm_iso_journey_key``.
ISO_JOURNEY_KEY_TO_FILENAME: dict[str, str] = {
    "iso_journey_heal_hospital_physician": "Iso Journey Heal Doctor.svg",
    "iso_journey_heal_long-term_resident": "Iso Journey Heal Senior.svg",
    "iso_journey_heal_outpatient_&_inpatient": "Iso Journey Heal Patient.svg",
    "iso_journey_heal_portfolio_manager": "Iso Journey Client Operation Director.svg",
    "iso_journey_heal_registered_nurse": "Iso Journey Heal Nurse.svg",
    "iso_journey_heal_site_manager": "Iso Journey Operator Site Manager.svg",
    "iso_journey_learn_middle_school_student": "Iso Journey Learn Schoolchild.svg",
    "iso_journey_learn_middle_school_teacher": "Iso Journey Learn Teacher.svg",
    "iso_journey_learn_portfolio_manager": "Iso Journey Client Operation Director.svg",
    "iso_journey_learn_site_manager": "Iso Journey Operator Site Manager.svg",
    "iso_journey_learn_university_student": "Iso Journey Learn Student.svg",
    "iso_journey_learn_working_parent_of_school-age_children": "Iso Journey Learn Parent.svg",
    "iso_journey_play_business_traveller_&_vip_guest": "Iso Journey Play VIP Guest Airport.svg",
    "iso_journey_play_cultural_tourist_&_museum_visitor": "Iso Journey Play Cultural Destination Visitor.svg",
    "iso_journey_play_executive_&_premium_event_guest": "Iso Journey Play VIP Guest Stadium.svg",
    "iso_journey_play_parent_&_sports_fan": "Iso Journey Play Football Fan.svg",
    "iso_journey_play_portfolio_manager": "Iso Journey Client Operation Director.svg",
    "iso_journey_play_site_manager": "Iso Journey Operator Site Manager.svg",
    "iso_journey_play_trade_show_&_event_attendent": "Iso Journey Play Event Participant.svg",
    "iso_journey_work_admin_function": "Iso Journey Work White Collar.svg",
    "iso_journey_work_army_officer": "Iso Journey Work Army Officer.svg",
    "iso_journey_work_lab_researcher_commute": "Iso Journey Work Grey Collar.svg",
    "iso_journey_work_portfolio_manager": "Iso Journey Client Operation Director.svg",
    "iso_journey_work_production_line_worker": "Iso Journey Work Blue Collar.svg",
    "iso_journey_work_site_manager": "Iso Journey Operator Site Manager.svg",
}

PILL_W, PILL_H = 16.0, 10.0


def norm_iso_journey_key(raw: Any) -> str:
    """Normalize Excel **Personae Iso Journey** cell for lookup (strip all whitespace, lower)."""
    return re.sub(r"\s+", "", str(raw or "").strip().lower())


def find_journey_asset(roots: list[Path], filename: str) -> Path | None:
    target = filename.lower()
    for root in roots:
        if not root.is_dir():
            continue
        direct = root / filename
        if direct.is_file():
            return direct
        for p in root.rglob("*"):
            if p.is_file() and p.name.lower() == target:
                return p
    return None


def norm_persona(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s).replace("\n", " ")).strip()


def include_row_for_persona(persona_id: str, excel_persona_raw: Any) -> bool:
    x = norm_persona(excel_persona_raw).lower()
    if persona_id == "tourist":
        return "business traveller" in x and "cultural tourist" not in x
    return True


def resolve_pair(area_raw: Any, persona_raw: Any) -> str | None:
    area = str(area_raw).strip().lower()
    pers = norm_persona(persona_raw).lower()
    return PAIR_TO_ID.get((area, pers))


def slugify(title: str) -> str:
    s = re.sub(r"\s+", " ", str(title).strip()).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return (s[:56] if s else "moment").strip("-")


def make_step_id(persona_id: str, title: str, used: set[str]) -> str:
    base = f"{persona_id}__{slugify(title)}"
    if len(base) > 80:
        base = base[:80].rstrip("-")
    k = base
    n = 2
    while k in used:
        suf = f"-{n}"
        trim = max(0, 80 - len(suf))
        k = (base[:trim].rstrip("-") + suf) if len(base) + len(suf) > 80 else base + suf
        n += 1
    used.add(k)
    return k


def header_indices(row1: tuple[Any, ...]) -> dict[str, int]:
    h: dict[str, int] = {}
    for i, cell in enumerate(row1):
        if cell is None:
            continue
        key = re.sub(r"\s+", " ", str(cell).strip().lower())
        h[key] = i

    def pick(*needles: str) -> int:
        for k, idx in h.items():
            if all(n in k for n in needles):
                return idx
        raise SystemExit(f"missing column containing {needles}; have {list(h.keys())}")

    def pick_body() -> int:
        best_i, best_len = -1, -1
        for k, idx in h.items():
            if "description" in k and "moment" in k and "subtitle" not in k:
                if len(k) > best_len:
                    best_len = len(k)
                    best_i = idx
        if best_i < 0:
            raise SystemExit(f"missing body column; have {list(h.keys())}")
        return best_i

    def pick_iso_journey_col() -> int:
        for k, idx in h.items():
            lk = k.lower()
            if "personae" in lk and "iso" in lk and "journey" in lk:
                return idx
        raise SystemExit(f"missing Personae Iso Journey column; have {list(h.keys())}")

    return {
        "area": pick("areas"),
        "personae": pick("personae"),
        "portrait": pick("personae", "portrait"),
        "iso_journey": pick_iso_journey_col(),
        "moment_title": pick("moments", "title"),
        "subtitle": pick("description", "subtitle"),
        "body": pick_body(),
        "hero_key": pick("image", "left", "moment"),  # Excel "Image left moment" (K) — moment raster SoT
        "modules": pick("modules"),
    }


def parse_modules(cell: Any) -> list[str]:
    if cell is None:
        return []
    s = str(cell).replace("\n", " ").strip()
    if not s:
        return []
    parts = re.split(r"[,;]", s)
    return [p.strip() for p in parts if p.strip()]


def norm_file_stem(stem: str) -> str:
    s = stem.lower()
    s = re.sub(r"_+", "_", s)
    return s


def build_image_index(images_dir: Path) -> dict[str, Path]:
    index: dict[str, Path] = {}
    for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
        for p in images_dir.glob(ext):
            index[norm_file_stem(p.stem)] = p
    return index


def resolve_image_path(excel_key: str, index: dict[str, Path], images_dir: Path) -> Path | None:
    k = str(excel_key or "").replace("\n", "").replace("\r", "").strip()
    if not k:
        return None
    stem = norm_file_stem(Path(k).name if "/" not in k else k.split("/")[-1])
    if stem.endswith(".png"):
        stem = norm_file_stem(stem[: -4])
    if stem in index:
        return index[stem]
    for ext in (".png", ".jpg", ".jpeg", ".webp"):
        cand = images_dir / f"{stem}{ext}"
        if cand.is_file():
            return cand
    candidates = sorted(index.items(), key=lambda x: SequenceMatcher(None, stem, x[0]).ratio(), reverse=True)
    if candidates and SequenceMatcher(None, stem, candidates[0][0]).ratio() >= 0.92:
        return candidates[0][1]
    return None


def ts_escape(s: str) -> str:
    return (
        s.replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\r\n", "\n")
        .replace("\r", "\n")
    )


def shrink_tile(path: Path, max_edge: int = 512) -> None:
    if platform.system() != "Darwin":
        return
    if path.suffix.lower() not in (".png", ".jpg", ".jpeg"):
        return
    try:
        subprocess.run(
            ["sips", "-Z", str(max_edge), str(path), "--out", str(path)],
            check=True,
            capture_output=True,
        )
    except (FileNotFoundError, subprocess.CalledProcessError):
        pass


def emit_journey_steps(steps: dict[str, dict[str, Any]]) -> str:
    lines = [
        "/** Auto-generated by scripts/ingest_personae_journey_excel.py — do not edit by hand. */",
        "",
        "import type { JourneyStep } from './types';",
        "",
        "export const EXCEL_JOURNEY_STEP_IDS = new Set<string>([",
    ]
    for sid in sorted(steps):
        lines.append(f'  "{sid}",')
    lines.append("]);")
    lines.append("")
    lines.append("export const JOURNEY_STEPS_FROM_EXCEL: Record<string, JourneyStep> = {")
    for sid in sorted(steps):
        s = steps[sid]
        mods = ", ".join(f'"{ts_escape(m)}"' for m in s["modules"])
        desc = ts_escape(str(s.get("description") or ""))
        lab = ts_escape(str(s["label"]))
        lines.append(f'  "{sid}": {{')
        lines.append(f'    id: "{sid}",')
        lines.append(f'    label: "{lab}",')
        lines.append(f'    icon: "📍",')
        lines.append(f'    modules: [{mods}],')
        lines.append(f'    description: "{desc}",')
        lines.append("  },")
    lines.append("};")
    lines.append("")
    return "\n".join(lines)


def emit_persona_journeys(journeys: dict[str, dict[str, Any]]) -> str:
    lines = [
        "/** Auto-generated by scripts/ingest_personae_journey_excel.py — do not edit by hand. */",
        "",
        "export const PERSONA_JOURNEYS_EXCEL: Record<",
        "  string,",
        "  {",
        "    image: string;",
        "    moments: Array<{ id: string; label: string; left: number; top: number; w: number; h: number }>;",
        "  }",
        "> = {",
    ]
    for pid in sorted(journeys):
        j = journeys[pid]
        lines.append(f'  "{pid}": {{')
        lines.append(f'    image: "{ts_escape(j["image"])}",')
        lines.append("    moments: [")
        for m in j["moments"]:
            lines.append(
                "      { "
                + f'id: "{m["id"]}", label: "{ts_escape(m["label"])}", '
                + f'left: {m["left"]}, top: {m["top"]}, w: {m["w"]}, h: {m["h"]} '
                + "},"
            )
        lines.append("    ],")
        lines.append("  },")
    lines.append("};")
    lines.append("")
    return "\n".join(lines)


def emit_editorial(ed: dict[str, dict[str, dict[str, str]]]) -> str:
    lines = [
        "/** Auto-generated by scripts/ingest_personae_journey_excel.py — do not edit by hand. */",
        "",
        "export interface MomentEditorialRow {",
        "  subtitle: string;",
        "  body: string;",
        "}",
        "",
        "export const MOMENT_EDITORIAL: Partial<",
        "  Record<string, Partial<Record<string, MomentEditorialRow>>>",
        "> = {",
    ]
    for pid in sorted(ed):
        lines.append(f'  "{pid}": {{')
        for sid in sorted(ed[pid]):
            row = ed[pid][sid]
            lines.append(f'    "{sid}": {{')
            lines.append(f'      subtitle: "{ts_escape(row["subtitle"])}",')
            lines.append(f'      body: "{ts_escape(row["body"])}",')
            lines.append("    },")
        lines.append("  },")
    lines.append("};")
    lines.append("")
    return "\n".join(lines)


def emit_raster(raster: dict[str, dict[str, str]]) -> str:
    lines = [
        "/** Auto-generated by scripts/ingest_personae_journey_excel.py — do not edit by hand. */",
        "",
        "export const MOMENT_HERO_RASTER: Partial<",
        "  Record<string, Partial<Record<string, string>>>",
        "> = {",
    ]
    for pid in sorted(raster):
        lines.append(f'  "{pid}": {{')
        for sid in sorted(raster[pid]):
            lines.append(f'    "{sid}": "{raster[pid][sid]}",')
        lines.append("  },")
    lines.append("};")
    lines.append("")
    return "\n".join(lines)


def emit_persona_top(top: dict[str, dict[str, str]]) -> str:
    lines = [
        "/** Auto-generated by scripts/ingest_personae_journey_excel.py — do not edit by hand. */",
        "",
        "export const MOMENT_PERSONA_TOP: Partial<",
        "  Record<string, Partial<Record<string, string>>>",
        "> = {",
    ]
    for pid in sorted(top):
        lines.append(f'  "{pid}": {{')
        for sid in sorted(top[pid]):
            lines.append(f'    "{sid}": "{top[pid][sid]}",')
        lines.append("  },")
    lines.append("};")
    lines.append("")
    lines.append(
        "export function momentPersonaTopUrl(personaId: string, stepId: string): string | undefined {"
    )
    lines.append("  return MOMENT_PERSONA_TOP[personaId]?.[stepId];")
    lines.append("}")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path.home() / "Downloads/Classeur Journey.xlsx")
    ap.add_argument(
        "--journey-assets",
        type=Path,
        action="append",
        default=[],
        help="Root folder for journey SVGs (recursive). May be passed multiple times.",
    )
    ap.add_argument(
        "--skip-moment-rasters",
        action="store_true",
        help="Do not require Journey_Moments_Images; skip hero tile copies.",
    )
    ap.add_argument(
        "--skip-portraits",
        action="store_true",
        help="Do not require Personae_Images_Portrait; skip persona-top copies.",
    )
    args = ap.parse_args()

    journey_roots = list(dict.fromkeys([p.expanduser() for p in args.journey_assets]))
    if not journey_roots:
        journey_roots = [Path.home() / "Downloads", Path.home() / "Downloads/moments"]

    moments_img = (
        Path.home()
        / "Library/CloudStorage/OneDrive-SODEXO/Design Community Hub - Documents"
        / "05_Sodexo Labs/Catalogue/Images catalogue/Journey_Moments_Images"
    )
    portrait_dir = (
        Path.home()
        / "Library/CloudStorage/OneDrive-SODEXO/Design Community Hub - Documents"
        / "05_Sodexo Labs/Catalogue/Images catalogue/Personae_Images_Portrait"
    )

    if not args.xlsx.exists():
        print(f"error: xlsx not found: {args.xlsx}", file=sys.stderr)
        return 2

    if not args.skip_moment_rasters and not moments_img.is_dir():
        print(f"error: Journey_Moments_Images not found: {moments_img}", file=sys.stderr)
        return 2
    if not args.skip_portraits and not portrait_dir.is_dir():
        print(f"error: portrait dir not found: {portrait_dir}", file=sys.stderr)
        return 2

    hero_index = build_image_index(moments_img) if moments_img.is_dir() else {}
    portrait_index = build_image_index(portrait_dir) if portrait_dir.is_dir() else {}

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb["Personae Journey"]
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col = header_indices(row1)

    ordered: dict[str, list[dict[str, Any]]] = defaultdict(list)
    skipped: list[tuple[str, str]] = []
    iso_carry: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col["area"]] is None:
            continue
        pid = resolve_pair(row[col["area"]], row[col["personae"]])
        if not pid:
            skipped.append((str(row[col["area"]]), norm_persona(row[col["personae"]])))
            continue
        if not include_row_for_persona(pid, row[col["personae"]]):
            continue
        title = row[col["moment_title"]]
        if title is None or not str(title).strip():
            continue
        iso_cell = row[col["iso_journey"]]
        iso_s = str(iso_cell or "").strip()
        if iso_s:
            iso_carry[pid] = iso_s
        elif pid in iso_carry:
            iso_s = iso_carry[pid]
        else:
            print(
                f"WARNING: moment row for persona {pid} missing Personae Iso Journey; skipped",
                file=sys.stderr,
            )
            continue
        ordered[pid].append(
            {
                "title": str(title).strip(),
                "modules": parse_modules(row[col["modules"]]),
                "subtitle": str(row[col["subtitle"]] or "").strip(),
                "body": str(row[col["body"]] or "").strip(),
                "hero_key": str(row[col["hero_key"]] or "").strip(),
                "portrait_key": str(row[col["portrait"]] or "").strip(),
                "iso_journey_key": iso_s,
            }
        )

    used_ids: set[str] = set()
    journeys: dict[str, dict[str, Any]] = {}
    journey_steps: dict[str, dict[str, Any]] = {}
    editorial: dict[str, dict[str, dict[str, str]]] = {}
    raster: dict[str, dict[str, str]] = {}
    top: dict[str, dict[str, str]] = {}
    persona_moments_out: dict[str, list[dict[str, str]]] = {}

    hero_out = REPO / "public/images/catalogue/assets/journeys/moments-raster"
    top_out = REPO / "public/images/catalogue/assets/journeys/moment-persona-top"
    journey_maps_out = REPO / "public/images/catalogue/assets/journeys/excel-maps"
    journey_maps_out.mkdir(parents=True, exist_ok=True)

    missing_hero: list[tuple[str, str, str]] = []
    missing_portrait: list[tuple[str, str, str]] = []

    for pid in sorted(ordered.keys()):
        rows = ordered[pid]
        keys_norm = {norm_iso_journey_key(r["iso_journey_key"]) for r in rows}
        keys_norm.discard("")
        if len(keys_norm) > 1:
            print(f"WARNING: persona {pid} has multiple Personae Iso Journey keys: {keys_norm}", file=sys.stderr)
        iso_key = next(iter(keys_norm), "")
        if not iso_key:
            print(f"error: persona {pid} has no Personae Iso Journey key", file=sys.stderr)
            return 2
        fname = ISO_JOURNEY_KEY_TO_FILENAME.get(iso_key)
        if not fname:
            print(
                f"error: unknown Personae Iso Journey key {rows[0]['iso_journey_key']!r} "
                f"(normalized {iso_key!r}) for persona {pid}",
                file=sys.stderr,
            )
            return 2
        src_map = find_journey_asset(journey_roots, fname)
        if not src_map:
            print(
                f"error: journey asset {fname!r} not found under {journey_roots} for persona {pid}",
                file=sys.stderr,
            )
            return 2

        ext = src_map.suffix.lower() or ".svg"
        dest_map = journey_maps_out / f"{pid}{ext}"
        shutil.copyfile(src_map, dest_map)
        map_url = f"/images/catalogue/assets/journeys/excel-maps/{pid}{ext}"

        with tempfile.TemporaryDirectory() as td:
            probe_png = Path(td) / "journey-probe.png"
            if not rasterize_journey_asset(src_map, probe_png):
                print(f"error: could not rasterize journey map {src_map} for {pid}", file=sys.stderr)
                return 2
            try:
                coords = hotspot_boxes_for_moments(probe_png, len(rows), PILL_W, PILL_H)
            except ValueError as e:
                print(f"error: hotspot detection for {pid}: {e}", file=sys.stderr)
                return 2

        moments_out: list[dict[str, Any]] = []
        pm_list: list[dict[str, str]] = []

        for i, r in enumerate(rows):
            sid = make_step_id(pid, r["title"], used_ids)
            left, top_pct, w, h = coords[i]
            moments_out.append(
                {
                    "id": sid,
                    "label": r["title"],
                    "left": round(left, 3),
                    "top": round(top_pct, 3),
                    "w": w,
                    "h": h,
                }
            )
            pm_list.append({"id": sid, "label": r["title"]})
            journey_steps[sid] = {
                "id": sid,
                "label": r["title"],
                "modules": r["modules"],
                "description": r["body"],
            }
            editorial.setdefault(pid, {})[sid] = {"subtitle": r["subtitle"], "body": r["body"]}

            if not args.skip_moment_rasters:
                src_h = resolve_image_path(r["hero_key"], hero_index, moments_img)
                if src_h:
                    dest_d = hero_out / pid
                    dest_d.mkdir(parents=True, exist_ok=True)
                    hext = src_h.suffix.lower() or ".png"
                    dest = dest_d / f"{sid}{hext}"
                    shutil.copyfile(src_h, dest)
                    shrink_tile(dest)
                    raster.setdefault(pid, {})[
                        sid
                    ] = f"/images/catalogue/assets/journeys/moments-raster/{pid}/{sid}{hext}"
                else:
                    missing_hero.append((pid, sid, r["hero_key"]))

            if not args.skip_portraits:
                src_p = resolve_image_path(r["portrait_key"], portrait_index, portrait_dir)
                if src_p:
                    dest_d = top_out / pid
                    dest_d.mkdir(parents=True, exist_ok=True)
                    pext = src_p.suffix.lower() or ".png"
                    dest = dest_d / f"{sid}{pext}"
                    shutil.copyfile(src_p, dest)
                    shrink_tile(dest)
                    top.setdefault(pid, {})[
                        sid
                    ] = f"/images/catalogue/assets/journeys/moment-persona-top/{pid}/{sid}{pext}"
                else:
                    missing_portrait.append((pid, sid, r["portrait_key"]))

        journeys[pid] = {"image": map_url, "moments": moments_out}
        persona_moments_out[pid] = pm_list

    (REPO / "src/lib/data/personaeJourneyExcel.generated.ts").write_text(
        emit_persona_journeys(journeys), encoding="utf-8"
    )
    (REPO / "src/lib/data/journeyStepsFromExcel.generated.ts").write_text(
        emit_journey_steps(journey_steps), encoding="utf-8"
    )
    (REPO / "src/lib/data/momentEditorial.generated.ts").write_text(emit_editorial(editorial), encoding="utf-8")
    if not args.skip_moment_rasters:
        (REPO / "src/lib/data/momentHeroRaster.generated.ts").write_text(
            emit_raster(raster), encoding="utf-8"
        )
    if not args.skip_portraits:
        (REPO / "src/lib/data/momentPersonaTop.generated.ts").write_text(
            emit_persona_top(top), encoding="utf-8"
        )
    (REPO / "scripts/persona-moments.json").write_text(
        json.dumps(persona_moments_out, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    print(f"personas: {len(journeys)}  steps: {len(journey_steps)}  hero tiles: {sum(len(v) for v in raster.values())}")
    if skipped:
        print(f"WARNING skipped unknown pairs: {len(skipped)}", skipped[:6])
    if missing_hero:
        print(f"WARNING missing hero ({len(missing_hero)}):", missing_hero[:5])
    if missing_portrait:
        print(f"WARNING missing portrait ({len(missing_portrait)}):", missing_portrait[:5])
    ok = not missing_hero and not missing_portrait
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
