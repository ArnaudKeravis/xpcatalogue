#!/usr/bin/env python3
"""Emit src/lib/data/solutionsExcelSoT.generated.ts from Classeur Solutions.xlsx (SoT)."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("openpyxl is required: pip install openpyxl") from e

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src/lib/data/solutionsExcelSoT.generated.ts"
DEFAULT_XLSX = Path.home() / "Downloads/Classeur Solutions.xlsx"
DEFAULT_MODULES_XLSX = Path.home() / "Downloads/Classeur Modules.xlsx"
SOLUTION_IMAGES_DIR = ROOT / "public/images/catalogue/assets/solutions"

IMAGE_EXT_ORDER = (".png", ".webp", ".jpg", ".jpeg", ".gif")


def esc(s: str) -> str:
    return json.dumps(s, ensure_ascii=True)


def norm_stem(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def build_image_index(img_dir: Path) -> dict[str, str]:
    """norm_stem(filename) -> chosen public URL path (/images/...). Prefers .png on ties."""
    buckets: dict[str, list[tuple[int, str, str]]] = {}
    if not img_dir.is_dir():
        return {}
    for p in img_dir.iterdir():
        if p.suffix.lower() not in IMAGE_EXT_ORDER:
            continue
        key = norm_stem(p.stem)
        pref = IMAGE_EXT_ORDER.index(p.suffix.lower()) if p.suffix.lower() in IMAGE_EXT_ORDER else 99
        buckets.setdefault(key, []).append((pref, p.name, p.suffix.lower()))
    out: dict[str, str] = {}
    for key, entries in buckets.items():
        entries.sort(key=lambda t: (t[0], t[1]))
        chosen = entries[0][1]
        out[key] = f"/images/catalogue/assets/solutions/{chosen}"
    return out


def hero_url(images_index: dict[str, str], solutions_images_key: str | None) -> str | None:
    if not solutions_images_key or not str(solutions_images_key).strip():
        return None
    raw = str(solutions_images_key).strip()
    tail = re.sub(r"^solution_image_", "", raw, flags=re.I).strip()
    key = norm_stem(tail)
    if key in images_index:
        return images_index[key]
    # Filename stems sometimes differ slightly (e.g. Affluence → affluences.png).
    for alt in (key + "s", key[:-1] if len(key) > 1 and key.endswith("s") else ""):
        if alt and alt in images_index:
            return images_index[alt]
    return None


def parse_module_solution_tokens(raw: str | None) -> list[str]:
    """Same token rules as the app: split on comma or semicolon, trim, drop empty."""
    if not raw or not str(raw).strip():
        return []
    return [t.strip() for t in re.split(r"[,;]", str(raw)) if t.strip()]


def validate_module_tokens_against_solutions(
    modules_path: Path,
    solution_names: list[str],
) -> tuple[list[tuple[str, str]], dict[str, list[str]]]:
    """
    Returns (missing_pairs, hints_by_token).
    missing_pairs: (module_name, token) where token is not an exact match to any solution name.
    hints_by_token: token -> list of solution names with the same normalized key (letters/digits only).
    """
    sol_set = set(solution_names)
    sol_by_norm: dict[str, list[str]] = {}
    for n in solution_names:
        sol_by_norm.setdefault(norm_stem(n), []).append(n)

    missing: list[tuple[str, str]] = []
    wb = openpyxl.load_workbook(modules_path, read_only=True, data_only=True)
    if "Modules" not in wb.sheetnames:
        raise ValueError(f"No 'Modules' sheet in {modules_path}")
    ws = wb["Modules"]
    rows_in = list(ws.iter_rows(values_only=True))
    if not rows_in:
        return [], {}

    header = [str(c).strip() if c is not None else "" for c in rows_in[0]]
    try:
        col_modules = header.index("Solutions in the module")
    except ValueError:
        col_modules = 2

    for row in rows_in[1:]:
        mod_name = row[0] if len(row) > 0 else None
        if mod_name is None or str(mod_name).strip() == "":
            continue
        raw = row[col_modules] if col_modules < len(row) else ""
        for tok in parse_module_solution_tokens(str(raw) if raw is not None else ""):
            if tok not in sol_set:
                missing.append((str(mod_name).strip(), tok))

    hints: dict[str, list[str]] = {}
    distinct_tokens = sorted({t for _, t in missing}, key=str.lower)
    for tok in distinct_tokens:
        nk = norm_stem(tok)
        if nk in sol_by_norm:
            hints[tok] = list(dict.fromkeys(sol_by_norm[nk]))

    return missing, hints


def print_module_solution_mismatch_report(
    missing: list[tuple[str, str]],
    hints: dict[str, list[str]],
) -> None:
    if not missing:
        print("Modules ↔ Solutions: all module tokens match a Solution name exactly.")
        return

    by_tok: dict[str, list[str]] = defaultdict(list)
    for mod, tok in missing:
        by_tok[tok].append(mod)

    print()
    print("=== Modules ↔ Solutions token check (exact string match) ===")
    print(f"Missing matches: {len(missing)} references across {len(by_tok)} distinct tokens.")
    print("Fix Classeur Modules.xlsx so each token matches Name of the solution in Classeur Solutions.xlsx.")
    print()

    for tok in sorted(by_tok.keys(), key=str.lower):
        mods = by_tok[tok]
        print(f"• Token not found in Solutions sheet: {tok!r}")
        print(f"  Modules using it ({len(mods)}): {', '.join(mods)}")
        if tok in hints:
            print(f"  Same normalized spelling in Solutions (copy/paste one): {hints[tok]}")
        print()


def main() -> None:
    import sys

    p = argparse.ArgumentParser(
        description="Generate solutionsExcelSoT.generated.ts and optionally validate Modules sheet tokens.",
    )
    p.add_argument(
        "xlsx",
        nargs="?",
        default=str(DEFAULT_XLSX),
        help="Path to Classeur Solutions.xlsx",
    )
    p.add_argument(
        "--modules",
        metavar="PATH",
        default=str(DEFAULT_MODULES_XLSX),
        help=f"Classeur Modules.xlsx for cross-check (default: {DEFAULT_MODULES_XLSX})",
    )
    p.add_argument(
        "--skip-modules-check",
        action="store_true",
        help="Do not compare Modules sheet tokens to Solution names.",
    )
    p.add_argument(
        "--strict",
        action="store_true",
        help="Exit with code 1 if any module token does not match a solution name exactly.",
    )
    args = p.parse_args()
    path = Path(args.xlsx).expanduser()
    if not path.is_file():
        raise SystemExit(f"Missing workbook: {path}")

    images_index = build_image_index(SOLUTION_IMAGES_DIR)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Solutions" not in wb.sheetnames:
        raise SystemExit(f"No 'Solutions' sheet in {path}")
    ws = wb["Solutions"]
    rows_in = list(ws.iter_rows(values_only=True))
    if not rows_in:
        raise SystemExit("Empty sheet")

    header = [str(c).strip() if c is not None else "" for c in rows_in[0]]
    col = {name: header.index(name) for name in header if name in header}

    def idx(name: str, fallback: int) -> int:
        return col.get(name, fallback)

    i_name = idx("Name of the solution", 0)
    i_sol_img = idx("Solutions Images", 1)
    i_desc_sol = idx("Description Solution", 2)
    i_regions = idx("Regions and country", 3)
    i_context = idx("Context", 4)
    i_description = idx("Description", 5)
    i_deploy = idx("Deployment & KPIs", 6)
    i_benefits = idx("Benefits", 7)
    i_contacts = idx("Contacts", 8)
    i_tag = idx("Tag for catalogue", 9)
    i_hashtags = idx("Hashtags", 10)

    rows: list[dict[str, str | None]] = []
    missing_img: list[str] = []

    for row in rows_in[1:]:
        name = row[i_name] if i_name < len(row) else None
        if name is None or str(name).strip() == "":
            continue
        sol_img = row[i_sol_img] if i_sol_img < len(row) else None
        hero = hero_url(images_index, str(sol_img).strip() if sol_img else None)
        if sol_img and str(sol_img).strip() and hero is None:
            missing_img.append(str(name).strip())

        def cell(j: int) -> str | None:
            if j < 0 or j >= len(row):
                return None
            v = row[j]
            if v is None:
                return None
            return str(v).strip() or None

        rows.append(
            {
                "name": str(name).strip(),
                "solutionsImagesKey": cell(i_sol_img),
                "descriptionSolutionKey": cell(i_desc_sol),
                "regionsAndCountry": cell(i_regions),
                "context": cell(i_context) or "",
                "description": cell(i_description) or "",
                "deploymentKpis": cell(i_deploy) or "",
                "benefits": cell(i_benefits) or "",
                "contacts": cell(i_contacts) or "",
                "tagForCatalogue": cell(i_tag),
                "hashtagsRaw": cell(i_hashtags),
                "heroImageUrl": hero,
            }
        )

    lines = [
        "/**",
        " * Auto-generated from the Sodexo Solutions workbook — do not edit by hand.",
        f" * Source rows: {len(rows)}",
        " */",
        "",
        "export interface SolutionsExcelRow {",
        "  name: string;",
        "  solutionsImagesKey: string | null;",
        "  descriptionSolutionKey: string | null;",
        "  regionsAndCountry: string | null;",
        "  context: string;",
        "  description: string;",
        "  deploymentKpis: string;",
        "  benefits: string;",
        "  contacts: string;",
        "  tagForCatalogue: string | null;",
        "  hashtagsRaw: string | null;",
        "  heroImageUrl: string | null;",
        "}",
        "",
        "export const SOLUTIONS_EXCEL_SOT: readonly SolutionsExcelRow[] = [",
    ]
    for r in rows:
        hero_lit = "null" if r["heroImageUrl"] is None else esc(r["heroImageUrl"])
        lines.append(
            "  {"
            f"name: {esc(r['name'])}, "
            f"solutionsImagesKey: {esc(r['solutionsImagesKey']) if r['solutionsImagesKey'] else 'null'}, "
            f"descriptionSolutionKey: {esc(r['descriptionSolutionKey']) if r['descriptionSolutionKey'] else 'null'}, "
            f"regionsAndCountry: {esc(r['regionsAndCountry']) if r['regionsAndCountry'] else 'null'}, "
            f"context: {esc(r['context'])}, "
            f"description: {esc(r['description'])}, "
            f"deploymentKpis: {esc(r['deploymentKpis'])}, "
            f"benefits: {esc(r['benefits'])}, "
            f"contacts: {esc(r['contacts'])}, "
            f"tagForCatalogue: {esc(r['tagForCatalogue']) if r['tagForCatalogue'] else 'null'}, "
            f"hashtagsRaw: {esc(r['hashtagsRaw']) if r['hashtagsRaw'] else 'null'}, "
            f"heroImageUrl: {hero_lit}"
            "},"
        )
    lines.append("];")
    lines.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT} ({len(rows)} solutions)")
    if missing_img:
        print(f"Warning: {len(missing_img)} rows had Solutions Images keys with no matching file under public/")

    solution_names = [r["name"] for r in rows]
    modules_path = Path(args.modules).expanduser()
    if not args.skip_modules_check:
        if not modules_path.is_file():
            print(f"Warning: Modules workbook not found ({modules_path}); skipping Modules ↔ Solutions check.")
        else:
            try:
                miss, hints = validate_module_tokens_against_solutions(modules_path, solution_names)
                print_module_solution_mismatch_report(miss, hints)
                if args.strict and miss:
                    print("Strict mode: failing due to module↔solution token mismatches.", file=sys.stderr)
                    raise SystemExit(1)
            except ValueError as e:
                print(f"Warning: {e}; skipping Modules ↔ Solutions check.")

if __name__ == "__main__":
    main()
